#!/usr/bin/env python3
"""
Krato Minerador v2 — preço REAL via Keepa.

Fluxo: Gemini gera IDEIAS (keywords, sem preço) -> Keepa dá o preço real
(média 90 dias, Amazon UK) + BSR -> calcula margem real -> filtra >= 60% ->
grava um relatório HTML no Airtable.

Roda no GitHub Actions (o Make é bloqueado pelo Keepa com 502; o Actions passa).

Secrets (env vars): KEEPA_KEY, GEMINI_API_KEY, AIRTABLE_TOKEN
"""
import os
import json
import time
import gzip
import html
import smtplib
import urllib.request
import urllib.parse
import urllib.error
from email.message import EmailMessage
from datetime import datetime, timezone, timedelta

# ---------- Config ----------
KEEPA_KEY = os.environ["KEEPA_KEY"]
GEMINI_KEY = os.environ["GEMINI_API_KEY"]
AIRTABLE_TOKEN = os.environ["AIRTABLE_TOKEN"]

AIRTABLE_BASE = "appv3M1eSHNB7Voq4"
AIRTABLE_TABLE = "tblR2ZFwfJOFghveb"
F_DATA = "fld92cl3FO8zKPW6l"       # Data (DD/MM/YYYY)
F_SESSION = "fldG5W2UbWY3PGqGw"    # Sessao
F_RESULT = "fldNfGOrSKpAtPYCF"     # Resultado (HTML)
F_SESSID = "fldqadmkDBQPZiBOJ"     # Name / session id

# E-mail (Gmail SMTP com App Password) — só envia se as creds existirem
GMAIL_USER = os.environ.get("GMAIL_USER", "")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
EMAIL_TO = os.environ.get("EMAIL_TO", "wellbuenorider@gmail.com")

N_IDEAS = int(os.environ.get("N_IDEAS", "8"))
BSR_MAX = int(os.environ.get("BSR_MAX", "20000"))     # BSR acima disso = vende pouco
PRICE_MIN = float(os.environ.get("PRICE_MIN", "18"))  # abaixo disso a taxa FIXA do FBA mata a margem
PRICE_MAX = float(os.environ.get("PRICE_MAX", "60"))  # descarta outlier caro
NET_FLOOR = float(os.environ.get("NET_FLOOR", "20"))  # piso de lucro LÍQUIDO (%) — regra Krato
KEEPA_DOMAIN = "2"  # amazon.co.uk

# Modelo de taxas Amazon UK FBA (aprox., empresa NÃO registrada no VAT).
# Usado p/ calcular o teto de custo (RFQ) que garante o líquido-alvo.
REFERRAL = 0.15      # comissão Amazon
FBA_FIXED = 3.00     # taxa FBA fixa (item médio) — £, não %
PPC_ALLOW = 0.15     # reserva p/ anúncio
RETURNS = 0.04       # devoluções
VAT_ON_FEES = 0.20   # IVA sobre as taxas Amazon (não recuperável sem registro)


def cost_ceiling_for_net(price, net_pct):
    """Custo landed máximo p/ atingir net_pct% de lucro líquido nesse preço."""
    referral = price * REFERRAL
    vat = (referral + FBA_FIXED) * VAT_ON_FEES
    ppc = price * PPC_ALLOW
    ret = price * RETURNS
    return round(price - (net_pct / 100.0) * price - referral - FBA_FIXED - vat - ppc - ret, 2)

# Clichês saturados que o agente NUNCA deve propor (mineração premium Krato)
EXCLUSION = [
    "bamboo hairbrush", "bamboo toothbrush holder", "mesh produce bags",
    "jade roller", "gua sha", "resistance bands", "silicone baby bibs",
    "dog snuffle mat", "bamboo bath caddy", "portable blender",
    "makeup remover pads", "salt lamp", "thermal label rolls", "qr stands",
    "barcode scanner", "tumbler", "mug", "candle", "reusable coffee cup",
]

BRAND = dict(navy="#1a1a2e", gold="#C8A96E", muted="#8a8a8a", body="#2c2c2c", bg="#eef0f2")


def _http(req, timeout=60):
    """Faz a requisição com retentativa em erro transitório (5xx / timeout / rede)."""
    last = None
    for attempt in range(4):
        try:
            with urllib.request.urlopen(req, timeout=timeout) as r:
                raw = r.read()
            if raw[:2] == b"\x1f\x8b":
                raw = gzip.decompress(raw)
            return json.loads(raw.decode())
        except urllib.error.HTTPError as e:
            last = e
            if e.code < 500:   # 4xx = erro do cliente, nao adianta repetir
                raise
            print(f"  [retry] HTTP {e.code} (tentativa {attempt + 1}/4), aguardando...")
        except Exception as e:
            last = e
            print(f"  [retry] {type(e).__name__} (tentativa {attempt + 1}/4), aguardando...")
        time.sleep(6 * (attempt + 1))
    raise last


def _get(url, timeout=60):
    req = urllib.request.Request(url, headers={"User-Agent": "krato-miner", "Accept-Encoding": "gzip"})
    return _http(req, timeout)


def _post(url, payload, timeout=60):
    body = json.dumps(payload).encode()
    req = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json"})
    return _http(req, timeout)


# ---------- 1. Gemini: gerar ideias ----------
def gemini_ideas(n):
    """Rotacao diaria de keywords por categoria — cada dia varia, nao repete o mesmo produto/vendedor."""
    SCANNER = [("wireless barcode scanner", "Leitor Sem Fio", "\U0001F4F6"), ("2d qr barcode scanner", "Leitor QR 2D", "\U0001F4F1"), ("bluetooth barcode scanner", "Leitor Bluetooth", "\U0001F535"), ("handheld barcode scanner", "Leitor de Mao", "\U0001F52B")]
    PRINTER = [("thermal label printer", "Impressora Termica", "\U0001F5A8"), ("4x6 shipping label printer", "Impressora 4x6", "\U0001F3F7"), ("bluetooth label printer", "Impressora Bluetooth", "\U0001F5A8")]
    PACKAGING = [("barcode label rolls", "Rolos de Etiqueta", "\U0001F3F7"), ("4x6 thermal labels", "Etiquetas 4x6", "\U0001F3F7"), ("jewelry hang tags", "Tags de Joia", "\U0001F516"), ("kraft swing tags", "Tags Kraft", "\U0001F516"), ("tissue paper wrapping", "Papel de Seda", "\U0001F381"), ("velvet jewelry pouch", "Saquinho de Veludo", "\U0001F49C"), ("poly mailer bags", "Envelope Plastico", "\U0001F4E6")]
    FASHION = [("western plaid shirt women", "Camisa Xadrez Western", "\U0001F457"), ("cowboy boots women", "Bota Cowboy Fem", "\U0001F462"), ("cowboy hat", "Chapeu Cowboy", "\U0001F920"), ("cowgirl boots", "Bota Cowgirl", "\U0001F462"), ("fringe suede jacket", "Jaqueta de Franjas", "\U0001F9E5"), ("western denim shirt men", "Camisa Jeans Western", "\U0001F454"), ("bandana", "Bandana", "\U0001F53B")]
    JEWELRY = [("gold hoop earrings", "Brinco de Argola", "\U0001F442"), ("layered necklace", "Colar em Camadas", "\U0001F4FF"), ("stainless steel ring", "Anel de Aco", "\U0001F48D"), ("pearl drop earrings", "Brinco de Perola", "\U0001F9AA"), ("stud earrings set", "Kit de Brincos", "✨"), ("pendant necklace", "Colar com Pingente", "\U0001F4FF"), ("charm bracelet", "Pulseira Berloque", "\U0001F517"), ("anklet", "Tornozeleira", "\U0001F9B6"), ("choker necklace", "Choker", "\U0001F380"), ("cuff bracelet", "Bracelete", "⭕"), ("huggie earrings", "Argolinha", "\U0001F442")]
    ACCESSORY = [("leather belt women", "Cinto de Couro", "\U0001FAA2"), ("silk scarf", "Lenco de Seda", "\U0001F9E3"), ("hair claw clip", "Presilha", "\U0001F487"), ("sunglasses women", "Oculos de Sol", "\U0001F576"), ("crossbody bag", "Bolsa Transversal", "\U0001F45C"), ("beaded bracelet", "Pulseira Micanga", "\U0001F4FF"), ("brooch pin", "Broche", "\U0001F4CC")]
    slots = [("hardware", SCANNER, 0), ("hardware", PRINTER, 3), ("packaging", PACKAGING, 1),
             ("country_fashion", FASHION, 2), ("jewelry", JEWELRY, 5), ("country_fashion", ACCESSORY, 4)]
    doy = datetime.now(timezone.utc).timetuple().tm_yday
    ideas = []
    for cat, pool, off in slots:
        en, pt, emoji = pool[(doy + off) % len(pool)]
        ideas.append({"name_en": en.title(), "name_pt": pt, "emoji": emoji,
                      "category": cat, "keyword": en, "country": "A confirmar", "supplier": "A confirmar"})
    return ideas[:n]


# ---------- 2. Keepa: preço real + BSR ----------
_keepa_tokens = [999]  # cache do último tokensLeft conhecido


def _keepa_pace(cost):
    """Espera se faltam tokens (plano 1 token/min)."""
    if _keepa_tokens[0] < cost + 2:
        wait = max(1, cost + 2 - _keepa_tokens[0]) * 61
        print(f"  [pacing] tokens baixos ({_keepa_tokens[0]}), aguardando {wait}s...")
        time.sleep(wait)
        _keepa_tokens[0] += wait // 60


def keepa_search_asins(keyword, n=5):
    _keepa_pace(10)
    term = urllib.parse.quote_plus(keyword)
    url = f"https://api.keepa.com/search?key={KEEPA_KEY}&domain={KEEPA_DOMAIN}&type=product&asins-only=1&term={term}"
    d = _get(url)
    if "tokensLeft" in d:
        _keepa_tokens[0] = d["tokensLeft"]
    return (d.get("asinList") or [])[:n]


def keepa_products(asins):
    """Um único /product com vários ASINs (1 token cada). Retorna preço+BSR de cada."""
    if not asins:
        return []
    _keepa_pace(len(asins) + 1)
    csv = ",".join(asins)
    url = f"https://api.keepa.com/product?key={KEEPA_KEY}&domain={KEEPA_DOMAIN}&stats=90&history=0&asin={csv}"
    d = _get(url)
    if "tokensLeft" in d:
        _keepa_tokens[0] = d["tokensLeft"]
    out = []
    for p in (d.get("products") or []):
        st = p.get("stats") or {}
        avg90 = st.get("avg90") or []

        def val(i):
            return avg90[i] if len(avg90) > i and avg90[i] not in (None, -1) else None
        price = val(1) or val(0)   # New price senão Amazon price (pence)
        imgs = p.get("images") or []
        img_name = (imgs[0].get("m") or imgs[0].get("l")) if imgs else ""
        image = f"https://m.media-amazon.com/images/I/{img_name}" if img_name else ""
        out.append({
            "asin": p.get("asin"),
            "price_gbp": round(price / 100, 2) if price else None,
            "bsr": val(3),
            "title": p.get("title", ""),
            "image": image,
        })
    return out


# ---------- 3. Montar relatório ----------
def enrich(idea):
    kw = idea.get("keyword", "").strip()
    out = dict(idea)
    if not kw:
        out["status"] = "sem keyword"
        return out
    try:
        asins = keepa_search_asins(kw, 3)
    except Exception as e:
        out["status"] = f"erro busca: {e}"
        return out
    if not asins:
        out["status"] = "nao encontrado na Amazon UK"
        return out
    try:
        prods = keepa_products(asins)
    except Exception as e:
        out["status"] = f"erro preco: {e}"
        return out
    # Entre os top resultados, escolhe o que VENDE mais (menor BSR) dentro da
    # faixa de preço razoável — evita a listagem outlier cara/atípica.
    cands = [p for p in prods if p["price_gbp"] and PRICE_MIN <= p["price_gbp"] <= PRICE_MAX and p["bsr"]]
    if not cands:
        out["status"] = "sem match vendavel na faixa"
        return out
    # Relevancia: o titulo tem que conter uma palavra-chave forte da busca
    # (evita casar "camisa xadrez" com um colete so porque vende mais).
    STOP = {"women", "womens", "men", "mens", "kids", "unisex", "adult", "with", "for"}
    kw_words = [w for w in kw.lower().split() if len(w) > 3 and w not in STOP]
    rel = [p for p in cands if any(w in (p.get("title") or "").lower() for w in kw_words)]
    if rel:
        cands = rel
    best = min(cands, key=lambda p: p["bsr"])
    out.update(best)
    venda = best["price_gbp"]
    out["cost_20"] = cost_ceiling_for_net(venda, 20)  # teto custo p/ 20% líquido (piso)
    out["cost_30"] = cost_ceiling_for_net(venda, 30)  # teto custo p/ 30% líquido (meta)
    out["status"] = "ok"
    return out


def card_html(x):
    e = lambda s: html.escape(str(s))
    bsr = x.get("bsr")
    hot = " 🔥" if bsr and bsr < 2000 else ""  # BSR baixo = vende muito
    amz = "https://www.amazon.co.uk/s?k=" + urllib.parse.quote_plus(x.get("name_en", ""))
    return f"""
<table width="100%" cellpadding="14" cellspacing="0" border="0" bgcolor="#ffffff"
 style="background-color:#ffffff;border:1px solid #eeeeee;margin-bottom:12px;font-family:sans-serif">
<tr><td>
  <img src="{e(x.get('image',''))}" alt="" width="100" style="float:right;margin:0 0 8px 12px;border:1px solid #eeeeee;background:#ffffff">
  <span style="background-color:{BRAND['gold']};color:#fff;font-size:10px;font-weight:bold;padding:2px 6px">{e(CAT_LABEL.get(x.get('category'), x.get('category','')))}</span><br>
  <span style="font-size:16px;font-weight:bold;color:{BRAND['navy']}">{e(x.get('name_en'))}{hot}</span>
  <span style="font-size:13px;color:{BRAND['muted']}"> {e(x.get('name_pt'))}</span> {e(x.get('emoji',''))}
  <br><br>
  <span style="color:{BRAND['muted']}">Preço REAL Amazon UK (Keepa 90d):</span> <b style="color:{BRAND['navy']};font-size:15px">£{e(x.get('price_gbp'))}</b>
  &nbsp;·&nbsp; <span style="color:{BRAND['muted']}">Demanda (BSR):</span> <b>{e(bsr or 'n/d')}</b>
  <br>
  <span style="color:{BRAND['gold']};font-weight:bold">🎯 Alvo de sourcing (RFQ): custo ≤ £{e(x.get('cost_20'))} p/ 20% líq · ≤ £{e(x.get('cost_30'))} p/ 30% líq</span>
  <span style="color:{BRAND['muted']};font-size:11px"> (já descontadas taxas Amazon + PPC)</span>
  <br>
  <span style="color:{BRAND['muted']};font-size:12px">Origem sugerida: {e(x.get('country'))} &nbsp;·&nbsp; Fornecedor: {e(x.get('supplier'))}</span>
  <br><span style="color:{BRAND['muted']};font-size:11px">Match Keepa: {e((x.get('title') or '')[:90])}</span>
  <br><br>
  <a href="{amz}" style="background-color:{BRAND['gold']};color:#fff;font-weight:bold;padding:8px 16px;text-decoration:none;font-size:12px">Ver na Amazon UK →</a>
</td></tr></table>"""


CAT_LABEL = {"country_fashion": "Moda Country", "jewelry": "Semijoia/Acessorio",
             "packaging": "Insumo/Embalagem", "hardware": "Hardware POS"}


def build_report(rows, date_str):
    # MODO PESQUISA: mostra o panorama real (preço + demanda) das 4 categorias
    # da plataforma p/ analisarmos. Sem filtro duro — queremos ver o mercado.
    passed = [r for r in rows if r.get("status") == "ok" and r.get("price_gbp") and r.get("bsr")]
    cat_order = {"country_fashion": 0, "jewelry": 1, "packaging": 2, "hardware": 3}
    passed.sort(key=lambda r: (cat_order.get(r.get("category"), 9), r.get("bsr") or 9e18))
    skipped = [r for r in rows if r not in passed]
    cards = "".join(card_html(r) for r in passed) or \
        f"<p style='color:{BRAND['muted']}'>Nenhum produto verificado bateu a margem real hoje.</p>"
    note = ""
    if skipped:
        items = "; ".join(f"{html.escape(r.get('name_en','?'))} ({html.escape(r.get('status','?'))})" for r in skipped)
        note = f"<p style='color:{BRAND['muted']};font-size:11px'>Descartados/não verificados: {items}</p>"
    return f"""<html><body style="background:{BRAND['bg']};padding:16px;font-family:sans-serif">
<table width="100%" style="max-width:640px" cellpadding="0" cellspacing="0" border="0" align="center"><tr><td>
<table width="100%" bgcolor="{BRAND['navy']}" style="background-color:{BRAND['navy']}" cellpadding="20"><tr><td>
  <span style="color:{BRAND['gold']};font-size:20px;font-weight:bold">Krato — Pesquisa de Mercado (Keepa)</span><br>
  <span style="color:{BRAND['muted']};font-size:13px">{date_str} · {len(passed)} produtos · Moda Country · Semijoia · Insumos · Hardware</span><br>
  <span style="color:{BRAND['gold']};font-size:11px">Preço + demanda = Amazon UK real (Keepa). Para análise de catálogo e sourcing.</span>
</td></tr></table>
<br>{cards}{note}
<p style="color:{BRAND['muted']};font-size:11px;border-left:3px solid {BRAND['gold']};padding-left:8px">KRATO GLOBAL · Well Bueno Limited · Uso interno · Confidencial</p>
</td></tr></table></body></html>"""


# ---------- 4. Airtable ----------
def airtable_write(date_str, sessid, html_report):
    url = f"https://api.airtable.com/v0/{AIRTABLE_BASE}/{AIRTABLE_TABLE}"
    payload = {"fields": {F_DATA: date_str, F_SESSION: "V2-KEEPA",
                          F_RESULT: html_report, F_SESSID: sessid}, "typecast": True}
    body = json.dumps(payload).encode()
    req = urllib.request.Request(url, data=body, method="POST",
                                 headers={"Authorization": f"Bearer {AIRTABLE_TOKEN}",
                                          "Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read().decode())


# ---------- 5. Planilha de leads (o que o Keepa entrega) ----------
def build_catalog_xlsx(rows):
    """Gera a planilha de leads do dia. Retorna (bytes, nome) ou (None, None) se faltar openpyxl."""
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception as e:
        print("  [xlsx] openpyxl indisponivel, e-mail sem anexo:", e)
        return None, None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    hdr = ["Produto", "PT", "Categoria", "Keyword (busca)", "Preco Amazon (GBP)",
           "BSR (demanda)", "Custo alvo 20% liq", "Custo alvo 30% liq", "Status", "Imagem"]
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A1A2E")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for r in rows:
        ws.append([r.get("name_en"), r.get("name_pt"), r.get("category"), r.get("keyword"),
                   r.get("price_gbp"), r.get("bsr"), r.get("cost_20"), r.get("cost_30"),
                   r.get("status"), r.get("image")])
    for row in ws.iter_rows(min_row=2):
        for i in (5, 7, 8):
            if isinstance(row[i - 1].value, (int, float)):
                row[i - 1].number_format = "£#,##0.00"
    for i, w in enumerate([22, 20, 15, 22, 16, 12, 16, 16, 22, 42], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), "krato-leads.xlsx"


# ---------- 6. E-mail (visual, com imagens + planilha anexa) ----------
def send_email(subject, html_report, xlsx_bytes=None, xlsx_name="krato-leads.xlsx"):
    if not (GMAIL_USER and GMAIL_APP_PASSWORD):
        print("  [email] sem creds Gmail - pulando envio (grava so no Airtable)")
        return
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = GMAIL_USER
    msg["To"] = EMAIL_TO
    msg.set_content("Seu leitor nao suporta HTML. Veja o relatorio no Airtable.")
    msg.add_alternative(html_report, subtype="html")
    if xlsx_bytes:
        msg.add_attachment(xlsx_bytes, maintype="application",
                           subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           filename=xlsx_name)
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=30) as s:
        s.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        s.send_message(msg)
    print("  [email] enviado para", EMAIL_TO, ("(com planilha)" if xlsx_bytes else "(sem anexo)"))


def main():
    now = datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=1)))  # ~UK
    date_str = now.strftime("%d/%m/%Y")
    sessid = now.strftime("%Y%m%d%H%M") + "-V2"
    print(f"== Krato Miner v2 == {date_str} | {N_IDEAS} ideias | BSR max {BSR_MAX} | faixa £{PRICE_MIN}-{PRICE_MAX}")

    ideas = gemini_ideas(N_IDEAS)
    print(f"Ideias do dia ({len(ideas)}, rotação):", [i.get("keyword") for i in ideas])

    rows = []
    for i, idea in enumerate(ideas, 1):
        r = enrich(idea)
        print(f"  {i}. {r.get('name_en')}: {r.get('status')}"
              + (f" | venda REAL £{r.get('price_gbp')} | BSR {r.get('bsr')} | custo<=£{r.get('cost_20')}(20% liq) / £{r.get('cost_30')}(30%)" if r.get("status") == "ok" else ""))
        rows.append(r)

    report = build_report(rows, date_str)
    res = airtable_write(date_str, sessid, report)
    print("Airtable record:", res.get("id"))
    xlsx_bytes, _ = build_catalog_xlsx(rows)
    send_email(f"Krato - Pesquisa de Mercado (Keepa) | {date_str}", report,
               xlsx_bytes, f"krato-leads-{now.strftime('%Y%m%d')}.xlsx")
    print("DONE")


if __name__ == "__main__":
    main()
